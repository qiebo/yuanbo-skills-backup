import json
import argparse
import shutil
import tempfile
import xml.etree.ElementTree as ET
from zipfile import ZIP_DEFLATED, ZipFile
from copy import deepcopy
from datetime import datetime, timezone, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path.cwd()
OUT = ROOT / "output"
EXCEL_METRICS = OUT / "excel_metrics.json"
WEBSITE_DATA = OUT / "website_data.json"
TEMPLATE_FIELDS = OUT / "template_fields.json"
XLSX_OUT = OUT / "yizhan-report-registry.xlsx"
JSON_OUT = OUT / "yizhan-report-registry.json"

TZ = timezone(timedelta(hours=8))
CREATED_AT = datetime.now(TZ).isoformat(timespec="seconds")


def load_json(path: Path):
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def pct_text(value):
    if value is None:
        return None
    return f"{value:.2%}"


def flatten_one_level(value):
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=False)


def safe_get(dct, *keys, default=None):
    cur = dct
    for key in keys:
        if isinstance(cur, dict) and key in cur:
            cur = cur[key]
        else:
            return default
    return cur


def append_row(rows, metric_id, name, value, unit, status, source_type, source_location, definition, formula="", note=""):
    rows.append(
        {
            "指标ID": metric_id,
            "指标名称": name,
            "数值": value,
            "单位": unit,
            "状态": status,
            "来源类型": source_type,
            "来源位置": source_location,
            "口径说明": definition,
            "公式或组件": formula,
            "备注": note,
        }
    )


def build_registry(excel, website, template):
    metrics = excel["metrics"]
    total_row = excel["total_row"]
    manual = website["manual_dom_aggregates"]
    api = website["api_aggregates"]
    audit = website.get("completeness_audit", {})
    trend = safe_get(audit, "website_full_list_monthly_record_trend", default={})
    trend_totals = trend.get("module_totals", {})

    student_accounts = manual["student_management_total_accounts"]
    total_logins = metrics["总登录人次"]["value"]
    coverage_rate = total_logins / student_accounts if student_accounts else None
    highest_module = metrics["功能模块总计比较"]["highest_modules"][0]
    highest_value = metrics["功能模块总计比较"]["highest_value"]
    bottle_stats = api["message_bottle_stats"]
    bottle_latest_total = safe_get(trend_totals, "message_bottle", "total_records")
    bottle_unique_sender = safe_get(trend_totals, "message_bottle", "unique_sender_count")

    core_rows = []
    append_row(
        core_rows,
        "CORE-001",
        "学校名称",
        excel["school_name"],
        "",
        "已获取",
        "Excel",
        "统计查询按月表：school_name",
        "目标学校名称，已与后台身份校验一致。",
    )
    append_row(
        core_rows,
        "CORE-002",
        "数据周期",
        f'{excel["data_period"]["start_month"]} 至 {excel["data_period"]["end_month"]}',
        "",
        "已获取",
        "Excel",
        "统计查询按月表：月度行",
        "当前报告覆盖的月份区间。",
    )
    append_row(
        core_rows,
        "CORE-003",
        "学生账号总数",
        student_accounts,
        "人",
        "已获取",
        "网站后台",
        "系统管理 > 学校信息管理 > 学生管理，分页总数",
        "覆盖率分母。",
    )
    append_row(
        core_rows,
        "CORE-004",
        "总登录人次",
        total_logins,
        "人次",
        "已获取",
        "Excel",
        "总计行：总登录人次",
        "覆盖率分子，按用户确认口径使用 Excel 总计行。",
    )
    append_row(
        core_rows,
        "CORE-005",
        "全校覆盖率",
        coverage_rate,
        "%",
        "公式生成",
        "Excel + 网站后台",
        "总登录人次 / 学生账号总数",
        "覆盖率 = 总登录人次 / 学生账号总数。",
        "CORE-004 / CORE-003",
    )
    append_row(
        core_rows,
        "CORE-006",
        "累计服务学生",
        metrics["累计服务学生"]["value"],
        "人",
        "已获取",
        "Excel",
        metrics["累计服务学生"]["source_location"],
        "沿用 Excel 总计行“总使用人数”口径。",
        note=metrics["累计服务学生"].get("note", ""),
    )
    append_row(
        core_rows,
        "CORE-007",
        "累计交互总量",
        metrics["累计交互总量"]["value"],
        "次",
        "已获取",
        "Excel/计算",
        metrics["累计交互总量"]["source_location"],
        "总计行中排除所有使用人数列后的功能使用数据求和。",
        metrics["累计交互总量"]["formula"],
        metrics["累计交互总量"].get("note", ""),
    )
    append_row(
        core_rows,
        "CORE-008",
        "本学期活跃峰值",
        metrics["本学期活跃峰值"]["value"],
        "人",
        "已获取",
        "Excel/计算",
        metrics["本学期活跃峰值"]["source_location"],
        "月度行“总使用人数”的最大值。",
        note=f'峰值月份：{metrics["本学期活跃峰值"]["peak_month"]}',
    )
    append_row(
        core_rows,
        "CORE-009",
        "平均登录次数",
        metrics["平均登录次数"]["value"],
        "次",
        "已获取",
        "Excel",
        metrics["平均登录次数"]["source_location"],
        "Excel 总计行原始字段。",
    )
    append_row(
        core_rows,
        "CORE-010",
        "全平台覆盖率最高模块",
        highest_module,
        "",
        "已获取",
        "Excel/计算",
        "总计行：按功能模块汇总使用数据",
        "对各模块总计数据比较，漂流瓶发送和回复合并计入漂流瓶。",
        json.dumps(metrics["功能模块总计比较"]["value"], ensure_ascii=False),
        f"最高值：{highest_value}",
    )
    append_row(
        core_rows,
        "CORE-011",
        "趋势数据可用口径",
        "后台记录数趋势",
        "",
        "部分可用",
        "网站后台/匿名本地聚合",
        "completeness_audit.website_full_list_monthly_record_trend",
        "该趋势为后台列表记录数趋势，不等同于使用次数趋势。",
        note="生成报告时如使用趋势，措辞必须写“后台记录数趋势”。",
    )

    module_rows = []
    def mod_row(module, metric, value, unit, status, source, location, definition, formula="", note=""):
        module_rows.append(
            {
                "模块": module,
                "指标": metric,
                "数值": value,
                "单位": unit,
                "状态": status,
                "来源类型": source,
                "来源位置": location,
                "口径说明": definition,
                "公式或组件": formula,
                "备注": note,
            }
        )

    career = manual["ai_career"]
    mod_row("AI生涯规划", "覆盖学生", career.get("total_users", 0), "人", "已获取", "网站后台", "AI生涯规划记录 > 互动记录", "网站汇总卡片：使用总人数。")
    mod_row("AI生涯规划", "累计使用次数", career.get("total_uses", 0), "次", "已获取", "网站后台", "AI生涯规划记录 > 互动记录", "网站汇总卡片：使用总次数。")
    mod_row("AI生涯规划", "本月使用次数", career.get("current_month_uses", 0), "次", "已获取", "网站后台", "AI生涯规划记录 > 互动记录", "网站汇总卡片：本月使用总次数。")
    mod_row("AI生涯规划", "后台列表记录数", career.get("table_total_rows", 0), "条", "已获取", "网站后台", "AI生涯规划记录 > 互动记录", "分页总数；与使用次数不同。")
    mod_row("AI生涯规划", "Excel功能使用量", total_row.get("ai_career_data_count", 0), "次", "已获取", "Excel", "总计行：AI生涯规划数据条数", "Excel导出功能使用量。")

    partner_history = manual.get("ai_partner_history", {})
    partner_content = api.get("partner_content_stats", [{}])[0] if api.get("partner_content_stats") else {}
    mod_row("AI心语伙伴", "累计对话", partner_history.get("total_uses", 0), "次", "已获取", "网站后台", "AI心语伙伴 > 历史分析", "历史分析口径，适合用于“累计对话”。")
    mod_row("AI心语伙伴", "覆盖学生", partner_history.get("total_users", 0), "人", "已获取", "网站后台", "AI心语伙伴 > 历史分析", "历史分析口径，适合用于“覆盖学生”。")
    mod_row("AI心语伙伴", "内容统计用户数", partner_content.get("user_count", 0), "人", "已获取", "网站后台", "AI心语伙伴 > 内容统计", "心理内容统计口径，不与历史分析覆盖学生混写。")
    mod_row("AI心语伙伴", "已分析人数", partner_content.get("analysis_count", 0), "人", "已获取", "网站后台", "AI心语伙伴 > 内容统计", "心理分析已覆盖人数。")
    mod_row("AI心语伙伴", "AI画像数", manual.get("ai_partner_profile", {}).get("table_total_rows", 0), "条", "已获取", "网站后台", "AI画像列表", "AI画像列表分页总数。")
    mod_row("AI心语伙伴", "Excel功能使用量", total_row.get("ai_partner_data_count", 0), "次", "已获取", "Excel", "总计行：AI心语伙伴数据条数", "Excel导出功能使用量。")

    bottle_delta_note = ""
    if bottle_latest_total and bottle_latest_total != bottle_stats.get("bottle_count", 0):
        bottle_delta_note = f"全量审计列表记录为 {bottle_latest_total}，与主聚合快照不同，可能是实时新增或口径差异。"
    mod_row("漂流瓶", "发布数量", bottle_stats.get("bottle_count", 0), "条", "已获取", "网站后台", "漂流瓶管理 > 发布记录 API", "主聚合快照中的发瓶数量。", note=bottle_delta_note)
    mod_row("漂流瓶", "回复数量", bottle_stats.get("reply_count", 0), "条", "已获取", "网站后台", "漂流瓶管理 > 发布记录 API", "主聚合快照中的回复数量。")
    mod_row("漂流瓶", "累计收发", bottle_stats.get("bottle_count", 0) + bottle_stats.get("reply_count", 0), "条", "公式生成", "网站后台/计算", "发布数量 + 回复数量", "漂流瓶累计收发量。", "发布数量 + 回复数量")
    mod_row("漂流瓶", "回复率", bottle_stats.get("reply_count", 0) / bottle_stats["bottle_count"] if bottle_stats.get("bottle_count") else None, "%", "公式生成", "网站后台/计算", "回复数量 / 发布数量", "以主聚合快照为准。", "回复数量 / 发布数量")
    mod_row("漂流瓶", "本月漂流瓶数量", bottle_stats.get("month_bottle_count", 0), "条", "已获取", "网站后台", "漂流瓶管理 > 发布记录 API", "主聚合快照中的本月发瓶数量。")
    mod_row("漂流瓶", "全量审计发瓶记录", bottle_latest_total, "条", "审计记录", "网站后台/匿名本地聚合", "website_full_list_monthly_record_trend.module_totals.message_bottle", "后续全量列表审计值，仅用于记录差异和趋势。")
    mod_row("漂流瓶", "发瓶学生去重数", bottle_unique_sender, "人", "条件可用", "网站后台/匿名本地聚合", "sender_id/user.id 去重", "仅代表发瓶学生去重；不包含回复参与者。")
    mod_row("漂流瓶", "Excel功能使用量", total_row.get("bottle_send_count", 0) + total_row.get("bottle_reply_count", 0), "次", "已获取", "Excel/计算", "总计行：漂流瓶发送数量 + 漂流瓶回复数量", "Excel功能模块比较口径。")

    interview = api.get("ai_interview_info", {})
    interview_reports = safe_get(api, "ai_interview_reports_meta", "numeric_total_like_fields", "data.pagination.total")
    mod_row("AI模拟面试", "服务学生", interview.get("total_users", 0), "人", "已获取", "网站后台", "AI模拟面试统计 API", "网站接口 total_users。")
    mod_row("AI模拟面试", "累计使用人次", interview.get("total_records", 0), "次", "已获取", "网站后台", "AI模拟面试统计 API", "网站接口 total_records。")
    mod_row("AI模拟面试", "本月使用人次", interview.get("current_month_records", 0), "次", "已获取", "网站后台", "AI模拟面试统计 API", "网站接口 current_month_records。")
    mod_row("AI模拟面试", "生成报告", interview_reports, "份", "已获取", "网站后台", "/api/ai/interview_stats/reports 分页总数", "报告列表总数。")
    mod_row("AI模拟面试", "深度使用率", interview_reports / interview["total_records"] if interview.get("total_records") else None, "%", "公式生成", "网站后台/计算", "生成报告 / 累计使用人次", "衡量从使用到报告生成的转化。", "生成报告 / 累计使用人次")
    mod_row("AI模拟面试", "Excel功能使用量", total_row.get("ai_interview_data_count", 0) + total_row.get("ai_report_count", 0), "次", "已获取", "Excel/计算", "总计行：AI面试数据条数 + AI输出报告数量", "Excel功能模块比较口径。")

    trend_rows = deepcopy(trend.get("rows", []))

    content_rows = []
    def content_row(module, dimension, item, category, value, unit, status, source, note=""):
        content_rows.append(
            {
                "模块": module,
                "维度": dimension,
                "对象": item,
                "分类或指标": category,
                "数值": value,
                "单位": unit,
                "状态": status,
                "来源位置": source,
                "备注": note,
            }
        )

    planner = api["planner_content_stats"]
    for idx, row in enumerate(planner.get("topQuestion", [[]])[0], start=1):
        status = "需筛选"
        note = "原始高频问题候选；正式报告只选用适合学校公开报告的表达。"
        if idx > 10:
            note += " 排名靠后或内容敏感时建议不进入正式报告。"
        content_row("AI生涯规划", "高频提问", idx, row.get("question"), row.get("count"), "次", status, "planner_content_stats.topQuestion", note)
    for idx, row in enumerate(planner.get("keywordList", [[]])[0][:30], start=1):
        content_row("AI生涯规划", "关键词", idx, row.get("keyword"), row.get("count"), "次", "已获取", "planner_content_stats.keywordList", "仅登记前30项，完整关键词仍保留在网站聚合 JSON。")
    for row in planner.get("tags", []):
        grade = row.get("grade")
        content_row("AI生涯规划", "年级使用占比", grade, "使用人数占比", row.get("use_rate"), "%", "已获取", "planner_content_stats.tags")
        for key, val in row.items():
            if key not in {"grade", "grade_id", "school_year", "use_rate"}:
                content_row("AI生涯规划", "年级互动内容分类", grade, key, val, "次", "已获取", "planner_content_stats.tags")

    for row in api.get("message_bottle_tag_stats", []):
        grade = row.get("grade")
        content_row("漂流瓶", "年级使用占比", grade, "使用人数占比", row.get("use_rate"), "%", "已获取", "message_bottle_tag_stats")
        for key, val in row.items():
            if key not in {"grade", "grade_id", "school_year", "use_rate"}:
                content_row("漂流瓶", "年级内容分类", grade, key, val, "条", "已获取", "message_bottle_tag_stats")

    for state, val in partner_content.get("analysis_stat", {}).items():
        public_name = {"无问题": "正常交流", "轻度": "初步关注", "中度": "关怀提醒", "重度": "特别关注"}.get(state, state)
        content_row("AI心语伙伴", "心理状态分级", public_name, state, val, "人", "已获取", "partner_content_stats.analysis_stat")
    for row in partner_content.get("analysis_grade", []):
        grade = row.get("grade_name")
        for key, val in row.items():
            if key != "grade_name":
                public_name = {"无问题": "正常交流", "轻度": "初步关注", "中度": "关怀提醒", "重度": "特别关注"}.get(key, key)
                content_row("AI心语伙伴", "心理状态年级分布", grade, public_name, val, "%", "已获取", "partner_content_stats.analysis_grade")
    for row in partner_content.get("advices_stat", {}).get("grade", []):
        grade = row.get("0")
        for key, val in row.items():
            if key != "0":
                content_row("AI心语伙伴", "心理困扰因子年级分布", grade, key, val, "%", "已获取", "partner_content_stats.advices_stat.grade")
    keywords_stat = partner_content.get("keywords_stat", {})
    if "top_events" in keywords_stat:
        for idx, row in enumerate(keywords_stat["top_events"], start=1):
            content_row("AI心语伙伴", "常问具体事件", idx, row.get("event"), row.get("占比"), "%", "已获取", "partner_content_stats.keywords_stat")
    else:
        for idx, (key, val) in enumerate(keywords_stat.items(), start=1):
            content_row("AI心语伙伴", "常问具体事件", idx, key, val, "%", "已获取", "partner_content_stats.keywords_stat")

    table_routes = []
    for row in audit.get("table_completeness", []):
        table_routes.append(
            {
                "模板表编号": row.get("table_no"),
                "用途": row.get("purpose"),
                "Task5状态": row.get("task5_status"),
                "推荐字段": flatten_one_level(row.get("recommended_fields")),
                "来源": flatten_one_level(row.get("source")),
                "备注": row.get("notes", ""),
            }
        )

    qa_rows = []
    def qa(item, status, evidence, action=""):
        qa_rows.append({"检查项": item, "状态": status, "证据": evidence, "后续动作": action})

    template_name = TEMPLATE_FIELDS.name if TEMPLATE_FIELDS else "未提供模板字段清单"
    qa("输入文件齐备", "通过", f"{EXCEL_METRICS.name}; {WEBSITE_DATA.name}; {template_name}")
    qa("学生账号总数", "通过", f"{student_accounts}，来自网站学生管理分页总数")
    qa("覆盖率公式", "通过", f"{total_logins} / {student_accounts} = {pct_text(coverage_rate)}")
    qa("累计交互总量口径", "通过", f"Excel 总计行排除所有使用人数列后的功能使用数据求和 = {metrics['累计交互总量']['value']}")
    qa("趋势数据口径", "限制", "已登记后台记录数趋势，不能写成使用次数趋势", "生成报告时使用谨慎措辞。")
    qa("漂流瓶实时差异", "限制", f"主聚合 {bottle_stats['bottle_count']}，后续全量审计 {bottle_latest_total}", "默认用主聚合快照，差异放备注。")
    qa("漂流瓶覆盖人数", "条件可用", f"发瓶学生去重 {bottle_unique_sender}", "若需要收发双向覆盖人数，标为未获取。")
    qa("AI心语伙伴多口径", "限制", f"历史分析 {partner_history['total_users']}/{partner_history['total_uses']}；内容统计 {partner_content.get('user_count')}/{partner_content.get('analysis_count')}；AI画像 {manual['ai_partner_profile']['table_total_rows']}", "根据报告表述选择口径，不混写。")
    qa("AI模拟面试生成报告", "通过", f"报告列表分页总数 {interview_reports}")
    qa("敏感信息持久化", "通过", "登记表仅保存聚合数据、字段来源和口径说明")

    return {
        "schema_version": "phase1-task6-data-registry-v1",
        "created_at": CREATED_AT,
        "source_files": {
            "excel_metrics": str(EXCEL_METRICS),
            "website_data": str(WEBSITE_DATA),
            "template_fields": str(TEMPLATE_FIELDS),
        },
        "school_name": excel["school_name"],
        "data_period": excel["data_period"],
        "core_metrics": core_rows,
        "module_metrics": module_rows,
        "website_backend_record_trend": trend_rows,
        "content_and_grade_stats": content_rows,
        "template_table_routes": table_routes,
        "qa_and_open_items": qa_rows,
    }


def setup_sheet(ws, widths):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(name="Microsoft YaHei", size=10)
            cell.border = Border(
                left=Side(style="thin", color="D9E2EC"),
                right=Side(style="thin", color="D9E2EC"),
                top=Side(style="thin", color="D9E2EC"),
                bottom=Side(style="thin", color="D9E2EC"),
            )
    for cell in ws[1]:
        cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.sheet_view.showGridLines = False


def write_table(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])


def write_registry_xlsx(registry):
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    module_lookup = {(row.get("模块"), row.get("指标")): row.get("数值") for row in registry.get("module_metrics", [])}
    bottle_count = module_lookup.get(("漂流瓶", "发布数量"), "未获取")
    bottle_reply = module_lookup.get(("漂流瓶", "回复数量"), "未获取")

    ws = wb.create_sheet("00_说明")
    info_rows = [
        ["项目", "内容"],
        ["产物", "生涯翼站报告数据登记表"],
        ["生成时间", registry["created_at"]],
        ["学校", registry["school_name"]],
        ["数据周期", f'{registry["data_period"]["start_month"]} 至 {registry["data_period"]["end_month"]}'],
        ["使用原则", "生成报告时只从本登记表取数；未获取/限制口径不得改写成确定结论。"],
        ["敏感信息策略", "仅登记聚合指标、来源路径、口径说明；不保存账号密码、cookie、token、学生个人明细。"],
        ["趋势口径", "当前可用趋势为网站后台记录数趋势，不等同于使用次数趋势。"],
        ["漂流瓶口径", f"主聚合快照为 {bottle_count}/{bottle_reply}；如全量审计记录不同，视为实时差异并记录。"],
        ["AI心语伙伴口径", "历史分析、内容统计、AI画像三个口径分开登记，生成报告时按文字表述选择。"],
    ]
    for row in info_rows:
        ws.append(row)
    setup_sheet(ws, [22, 110])

    headers = ["指标ID", "指标名称", "数值", "单位", "状态", "来源类型", "来源位置", "口径说明", "公式或组件", "备注"]
    ws = wb.create_sheet("01_核心指标")
    write_table(ws, headers, registry["core_metrics"])
    setup_sheet(ws, [14, 22, 18, 10, 12, 16, 34, 44, 34, 44])
    core_row_by_id = {ws.cell(row=i, column=1).value: i for i in range(2, ws.max_row + 1)}
    ws.cell(row=core_row_by_id["CORE-005"], column=3).value = f'=IFERROR(C{core_row_by_id["CORE-004"]}/C{core_row_by_id["CORE-003"]},0)'
    ws.cell(row=core_row_by_id["CORE-005"], column=3).number_format = "0.00%"

    module_headers = ["模块", "指标", "数值", "单位", "状态", "来源类型", "来源位置", "口径说明", "公式或组件", "备注"]
    ws = wb.create_sheet("02_模块指标")
    write_table(ws, module_headers, registry["module_metrics"])
    setup_sheet(ws, [16, 20, 16, 10, 12, 18, 38, 44, 32, 44])
    row_map = {(ws.cell(row=i, column=1).value, ws.cell(row=i, column=2).value): i for i in range(2, ws.max_row + 1)}
    formula_specs = {
        ("漂流瓶", "累计收发"): f'=IFERROR(C{row_map[("漂流瓶", "发布数量")]}+C{row_map[("漂流瓶", "回复数量")]},0)',
        ("漂流瓶", "回复率"): f'=IFERROR(C{row_map[("漂流瓶", "回复数量")]}/C{row_map[("漂流瓶", "发布数量")]},0)',
        ("AI模拟面试", "深度使用率"): f'=IFERROR(C{row_map[("AI模拟面试", "生成报告")]}/C{row_map[("AI模拟面试", "累计使用人次")]},0)',
    }
    for key, formula in formula_specs.items():
        row_num = row_map[key]
        ws.cell(row=row_num, column=3).value = formula
        if key[1].endswith("率"):
            ws.cell(row=row_num, column=3).number_format = "0.00%"

    trend_headers = ["month", "ai_career_dialog_records", "ai_partner_analysis_records", "message_bottle_posts", "ai_interview_reports", "total_backend_records"]
    ws = wb.create_sheet("03_网站记录趋势")
    write_table(ws, trend_headers, registry["website_backend_record_trend"])
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=6).value = f"=SUM(B{r}:E{r})"
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1).value = "合计"
    for c in range(2, 7):
        col = get_column_letter(c)
        ws.cell(row=total_row, column=c).value = f"=SUM({col}2:{col}{total_row-1})"
        ws.cell(row=total_row, column=c).font = Font(name="Microsoft YaHei", size=10, bold=True)
    setup_sheet(ws, [14, 24, 26, 24, 22, 22])

    content_headers = ["模块", "维度", "对象", "分类或指标", "数值", "单位", "状态", "来源位置", "备注"]
    ws = wb.create_sheet("04_内容与年级统计")
    write_table(ws, content_headers, registry["content_and_grade_stats"])
    setup_sheet(ws, [16, 22, 22, 44, 12, 10, 14, 34, 54])

    route_headers = ["模板表编号", "用途", "Task5状态", "推荐字段", "来源", "备注"]
    ws = wb.create_sheet("05_模板字段路由")
    write_table(ws, route_headers, registry["template_table_routes"])
    setup_sheet(ws, [12, 30, 36, 70, 34, 60])

    qa_headers = ["检查项", "状态", "证据", "后续动作"]
    ws = wb.create_sheet("06_QA与未决项")
    write_table(ws, qa_headers, registry["qa_and_open_items"])
    setup_sheet(ws, [28, 12, 70, 56])

    for ws in wb.worksheets:
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

    wb.save(XLSX_OUT)
    normalize_styles_xml(XLSX_OUT)


def normalize_styles_xml(path: Path):
    """Reorder font child nodes so strict OpenXML validators accept the file."""
    ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    ET.register_namespace("x", ns["x"])
    order = {
        "b": 1,
        "i": 2,
        "strike": 3,
        "condense": 4,
        "extend": 5,
        "outline": 6,
        "shadow": 7,
        "u": 8,
        "vertAlign": 9,
        "sz": 10,
        "color": 11,
        "name": 12,
        "family": 13,
        "scheme": 14,
        "charset": 15,
    }
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        with ZipFile(path, "r") as zin:
            zin.extractall(tmp_path)
        styles_path = tmp_path / "xl" / "styles.xml"
        tree = ET.parse(styles_path)
        root = tree.getroot()
        fonts = root.find("x:fonts", ns)
        if fonts is not None:
            for font in fonts.findall("x:font", ns):
                children = list(font)
                children.sort(key=lambda el: order.get(el.tag.split("}", 1)[-1], 999))
                for child in list(font):
                    font.remove(child)
                for child in children:
                    font.append(child)
        tree.write(styles_path, encoding="utf-8", xml_declaration=True)

        new_path = path.with_suffix(".tmp.xlsx")
        with ZipFile(new_path, "w", ZIP_DEFLATED) as zout:
            for item in tmp_path.rglob("*"):
                if item.is_file():
                    zout.write(item, item.relative_to(tmp_path).as_posix())
        shutil.move(str(new_path), path)


def verify_workbook(path: Path):
    wb = load_workbook(path, data_only=False)
    required_sheets = {
        "00_说明",
        "01_核心指标",
        "02_模块指标",
        "03_网站记录趋势",
        "04_内容与年级统计",
        "05_模板字段路由",
        "06_QA与未决项",
    }
    missing = sorted(required_sheets - set(wb.sheetnames))
    if missing:
        raise RuntimeError(f"Missing sheets: {missing}")

    forbidden = ["token", "authorization", "cookie", "password", "passwd", "身份证", "手机号"]
    hits = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = "" if cell.value is None else str(cell.value)
                lower = value.lower()
                if any(term in lower for term in forbidden):
                    if "不保存账号密码" in value or "敏感信息持久化" in value:
                        continue
                    hits.append((ws.title, cell.coordinate, value[:100]))
    if hits:
        raise RuntimeError(f"Potential sensitive terms found: {hits[:5]}")

    formula_cells = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_cells.append(f"{ws.title}!{cell.coordinate}")
    return {"sheet_count": len(wb.sheetnames), "formula_cells": formula_cells}


def main():
    global EXCEL_METRICS, WEBSITE_DATA, TEMPLATE_FIELDS, XLSX_OUT, JSON_OUT, OUT
    parser = argparse.ArgumentParser(description="Build the 生涯翼站 report registry from Excel metrics and website aggregates.")
    parser.add_argument("--excel", required=True, help="Path to excel_metrics.json from extract_excel_metrics.py.")
    parser.add_argument("--website", required=True, help="Path to website_data.json collected from the SaaS backend.")
    parser.add_argument("--template-fields", default=None, help="Optional template field inventory JSON. If omitted, route sheet is left empty.")
    parser.add_argument("--out-dir", required=True, help="Output directory.")
    parser.add_argument("--basename", default=None, help="Output basename without extension.")
    args = parser.parse_args()

    EXCEL_METRICS = Path(args.excel).resolve()
    WEBSITE_DATA = Path(args.website).resolve()
    TEMPLATE_FIELDS = Path(args.template_fields).resolve() if args.template_fields else None
    OUT = Path(args.out_dir).resolve()
    OUT.mkdir(parents=True, exist_ok=True)

    excel = load_json(EXCEL_METRICS)
    website = load_json(WEBSITE_DATA)
    template = load_json(TEMPLATE_FIELDS) if TEMPLATE_FIELDS else {"schema_version": "template-fields-omitted", "tables": []}
    basename = args.basename or f'{excel.get("school_name", "学校")}-生涯翼站报告数据汇总'
    XLSX_OUT = OUT / f"{basename}.xlsx"
    JSON_OUT = OUT / f"{basename}.json"
    registry = build_registry(excel, website, template)

    with JSON_OUT.open("w", encoding="utf-8") as f:
        json.dump(registry, f, ensure_ascii=False, indent=2)

    write_registry_xlsx(registry)
    verification = verify_workbook(XLSX_OUT)

    print(
        json.dumps(
            {
                "status": "ok",
                "xlsx": str(XLSX_OUT),
                "json": str(JSON_OUT),
                "core_rows": len(registry["core_metrics"]),
                "module_rows": len(registry["module_metrics"]),
                "trend_rows": len(registry["website_backend_record_trend"]),
                "content_rows": len(registry["content_and_grade_stats"]),
                "route_rows": len(registry["template_table_routes"]),
                "qa_rows": len(registry["qa_and_open_items"]),
                "verification": verification,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
