import argparse
import json
from datetime import date, datetime, timezone, timedelta
from pathlib import Path

from openpyxl import load_workbook


TZ = timezone(timedelta(hours=8))

EXPECTED_HEADERS = [
    "school_name",
    "month",
    "ai_career_data_count",
    "ai_career_user_count",
    "ai_partner_data_count",
    "ai_partner_user_count",
    "bottle_send_count",
    "bottle_reply_count",
    "bottle_user_count",
    "ai_interview_data_count",
    "ai_interview_user_count",
    "ai_report_count",
    "total_user_count",
    "total_login_count",
    "avg_login_count",
]

FUNCTIONAL_USAGE_COLUMNS = [
    "ai_career_data_count",
    "ai_partner_data_count",
    "bottle_send_count",
    "bottle_reply_count",
    "ai_interview_data_count",
    "ai_report_count",
]

USER_COUNT_COLUMNS = [
    "ai_career_user_count",
    "ai_partner_user_count",
    "bottle_user_count",
    "ai_interview_user_count",
    "total_user_count",
]


def cell_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def normalize_month(value):
    value = cell_value(value)
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return f"{value.year:04d}-{value.month:02d}"
    text = str(value).strip()
    if text == "总计":
        return text
    text = text.replace("年", "-").replace("月", "").replace("/", "-").replace(".", "-")
    parts = [p for p in text.split("-") if p]
    if len(parts) >= 2 and parts[0].isdigit() and parts[1].isdigit():
        return f"{int(parts[0]):04d}-{int(parts[1]):02d}"
    return text


def to_number(value):
    value = cell_value(value)
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).replace(",", "").replace("%", "").strip()
    try:
        number = float(text)
    except ValueError:
        return value
    return int(number) if number.is_integer() else number


def find_header_row(ws):
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = [cell_value(v) for v in row]
        joined = " ".join(str(v) for v in values if v is not None)
        if "月份" in joined and ("总登录人次" in joined or "平均登录次数" in joined):
            return row_idx, values
    raise RuntimeError("Cannot find header row containing 月份 and login columns.")


def extract_excel_metrics(xlsx_path, sheet_name=None):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    header_row, raw_headers = find_header_row(ws)
    rows = []
    for excel_row_idx in range(header_row + 1, ws.max_row + 1):
        raw_values = [cell_value(ws.cell(excel_row_idx, col).value) for col in range(1, len(EXPECTED_HEADERS) + 1)]
        if not any(v is not None for v in raw_values):
            continue
        row = {"excel_row_index": excel_row_idx}
        for key, value in zip(EXPECTED_HEADERS, raw_values):
            row[key] = normalize_month(value) if key == "month" else to_number(value)
        rows.append(row)

    total_rows = [r for r in rows if r.get("month") == "总计"]
    if not total_rows:
        raise RuntimeError("Cannot find 总计 row in monthly statistics sheet.")
    total_row = total_rows[-1]
    monthly_rows = [r for r in rows if r.get("month") and r.get("month") != "总计"]
    if not monthly_rows:
        raise RuntimeError("No monthly rows found before 总计 row.")

    school_name = next((r.get("school_name") for r in monthly_rows if r.get("school_name")), None)
    if not school_name:
        raise RuntimeError("Cannot find school name in monthly rows.")

    months = [r["month"] for r in monthly_rows]
    total_interactions = sum(total_row.get(col) or 0 for col in FUNCTIONAL_USAGE_COLUMNS)
    active_row = max(monthly_rows, key=lambda r: r.get("total_user_count") or 0)
    module_totals = {
        "AI生涯规划": total_row.get("ai_career_data_count") or 0,
        "AI心语伙伴": total_row.get("ai_partner_data_count") or 0,
        "漂流瓶": (total_row.get("bottle_send_count") or 0) + (total_row.get("bottle_reply_count") or 0),
        "AI模拟面试": (total_row.get("ai_interview_data_count") or 0) + (total_row.get("ai_report_count") or 0),
    }
    highest_value = max(module_totals.values()) if module_totals else None
    highest_modules = [k for k, v in module_totals.items() if v == highest_value]

    return {
        "schema_version": "phase1-excel-metrics-v1",
        "created_at": datetime.now(TZ).isoformat(timespec="seconds"),
        "source_file": str(Path(xlsx_path).resolve()),
        "sheet_name": ws.title,
        "raw_headers": raw_headers[: len(EXPECTED_HEADERS)],
        "normalized_headers": EXPECTED_HEADERS,
        "header_notes": {
            "duplicate_columns": "原表存在多个“使用人数”列，按固定列位规范化命名。",
            "functional_usage_columns": FUNCTIONAL_USAGE_COLUMNS,
            "excluded_user_count_columns": USER_COUNT_COLUMNS,
        },
        "school_name": school_name,
        "data_period": {
            "start_month": min(months),
            "end_month": max(months),
            "monthly_row_count": len(monthly_rows),
        },
        "raw_rows": rows,
        "total_row": total_row,
        "metrics": {
            "累计服务学生": {
                "value": total_row.get("total_user_count"),
                "source_type": "Excel",
                "source_location": "总计行：总使用人数",
                "note": "该字段沿用 Excel 总计行口径。",
            },
            "累计交互总量": {
                "value": total_interactions,
                "source_type": "Excel/计算",
                "source_location": "总计行：排除所有使用人数列后的功能使用数据列",
                "formula": " + ".join(FUNCTIONAL_USAGE_COLUMNS),
                "component_values": {col: total_row.get(col) for col in FUNCTIONAL_USAGE_COLUMNS},
                "note": "按固定工作流口径计算。",
            },
            "本学期活跃峰值": {
                "value": active_row.get("total_user_count"),
                "source_type": "Excel/计算",
                "source_location": "月度行：总使用人数列最大值",
                "peak_month": active_row.get("month"),
                "note": "本学期活跃按峰值口径。",
            },
            "总登录人次": {
                "value": total_row.get("total_login_count"),
                "source_type": "Excel",
                "source_location": "总计行：总登录人次",
                "note": "后续用于登录人次覆盖强度分子。",
            },
            "平均登录次数": {
                "value": total_row.get("avg_login_count"),
                "source_type": "Excel",
                "source_location": "总计行：平均登录次数",
            },
            "功能模块总计比较": {
                "value": module_totals,
                "source_type": "Excel/计算",
                "source_location": "总计行：按功能模块汇总使用数据",
                "highest_modules": highest_modules,
                "highest_value": highest_value,
                "note": "漂流瓶=发送数量+回复数量；AI模拟面试=AI面试数据条数+AI输出报告数量。",
            },
        },
        "trend_analysis_policy": {
            "source_type": "网站",
            "note": "趋势分析数据从网站后台统计获取，不通过 Excel 表格获取。",
        },
    }


def main():
    parser = argparse.ArgumentParser(description="Extract 生涯翼站 monthly Excel metrics.")
    parser.add_argument("--xlsx", required=True, help="Input 统计查询按月 .xlsx path.")
    parser.add_argument("--sheet", default=None, help="Optional sheet name; default first sheet.")
    parser.add_argument("--out", required=True, help="Output excel_metrics.json path.")
    args = parser.parse_args()
    data = extract_excel_metrics(args.xlsx, args.sheet)
    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(out)


if __name__ == "__main__":
    main()
