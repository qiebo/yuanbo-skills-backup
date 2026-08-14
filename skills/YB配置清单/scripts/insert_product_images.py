#!/usr/bin/env python3
"""Insert remote ERP product image references into a configuration list.

Image strategy B (references/universal-rules.md §1): product images are remote
ERP `image_url` references. They are never downloaded, never locally hashed, and
never embedded as local files. For every selected output row, the binding
manifest carries the catalog `product_key` (= ERP `product_no`); we resolve the
single primary remote image ref and write a hyperlink to the ERP `image_url`
into the 产品图片 column. Product-name-only matching is refused.
"""

from __future__ import annotations

import argparse
import json
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


ERP_IMAGE_MATCH_RULE = "erp-image-url-v1"


def primary_remote_url(product: dict) -> str | None:
    """Return the ERP `image_url` of the single primary remote image ref, else None.

    Only a ref with role=primary, confidence=remote, match_rule=erp-image-url-v1
    and locally_verified=false is accepted.
    """
    refs = [
        ref
        for ref in product.get("image_refs", [])
        if ref.get("role") == "primary"
        and ref.get("confidence") == "remote"
        and ref.get("match_rule") == ERP_IMAGE_MATCH_RULE
        and ref.get("locally_verified") is False
    ]
    if len(refs) != 1:
        return None
    return refs[0].get("url") or None


def image_column(sheet, header_row: int, header: str = "产品图片") -> int:
    for column in range(1, sheet.max_column + 1):
        if sheet.cell(header_row, column).value == header:
            return column
    column = sheet.max_column + 1
    anchor = sheet.cell(header_row, max(1, column - 1))
    target = sheet.cell(header_row, column, header)
    if anchor.has_style:
        target._style = copy(anchor._style)
        target.font = copy(anchor.font)
        target.fill = copy(anchor.fill)
        target.border = copy(anchor.border)
        target.number_format = anchor.number_format
        target.protection = copy(anchor.protection)
    target.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.column_dimensions[get_column_letter(column)].width = 17
    return column


def insert_images(
    *,
    input_path: Path,
    output_path: Path,
    catalog_path: Path,
    assets_root: Path | None = None,
    sheet_name: str,
    header_row: int,
    bindings: list[dict],
) -> dict:
    """Write remote ERP image hyperlinks into the output workbook.

    `assets_root` is accepted for backward compatibility but unused: image
    strategy B references remote URLs, not local files.
    """
    input_path = input_path.resolve()
    output_path = output_path.resolve()
    if input_path == output_path:
        raise ValueError("Input and output workbooks must be different files.")
    catalog = json.loads(catalog_path.read_text(encoding="utf-8"))
    products_by_key: dict[str, dict] = {}
    for product in catalog.get("products", []):
        product_key = product.get("product_key")
        if not product_key:
            continue
        if product_key in products_by_key:
            raise ValueError(f"Duplicate product_key in catalog: {product_key}")
        products_by_key[product_key] = product
    workbook = load_workbook(input_path)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Output worksheet was not found: {sheet_name}")
    sheet = workbook[sheet_name]
    column = image_column(sheet, header_row)
    seen_rows: set[int] = set()
    warnings: list[dict] = []
    inserted = 0
    for binding in bindings:
        row = binding.get("row")
        product_key = binding.get("product_key")
        if not isinstance(row, int) or row <= header_row:
            raise ValueError(f"Invalid output row in binding: {row}")
        if row in seen_rows:
            raise ValueError(f"Duplicate output row in binding manifest: {row}")
        seen_rows.add(row)
        product = products_by_key.get(product_key)
        if product is None:
            warnings.append({"row": row, "product_key": product_key, "reason": "unknown product_key"})
            continue
        url = primary_remote_url(product)
        if not url:
            warnings.append(
                {
                    "row": row,
                    "product_key": product_key,
                    "reason": "no single primary remote image ref (locally_verified=false, match_rule=erp-image-url-v1)",
                }
            )
            continue
        cell = sheet.cell(row, column)
        cell.value = "查看图片"
        cell.hyperlink = url
        cell.font = Font(color="0563C1", underline="single")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        current_height = sheet.row_dimensions[row].height or 15
        sheet.row_dimensions[row].height = max(current_height, 24)
        inserted += 1
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return {"inserted": inserted, "warnings": warnings, "image_column": get_column_letter(column)}


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input-xlsx", required=True, type=Path)
    parser.add_argument("--output-xlsx", required=True, type=Path)
    parser.add_argument("--catalog", required=True, type=Path)
    parser.add_argument("--assets-root", required=False, type=Path, default=None)
    parser.add_argument("--manifest", required=True, type=Path)
    args = parser.parse_args()
    manifest = json.loads(args.manifest.read_text(encoding="utf-8"))
    result = insert_images(
        input_path=args.input_xlsx,
        output_path=args.output_xlsx,
        catalog_path=args.catalog,
        assets_root=args.assets_root,
        sheet_name=manifest["sheet_name"],
        header_row=manifest["header_row"],
        bindings=manifest["bindings"],
    )
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
