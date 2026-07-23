import hashlib
import json
import os
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from PIL import Image

from scripts.insert_product_images import insert_images


class InsertProductImagesTests(unittest.TestCase):
    def test_skill_requires_unmatched_requirements_in_both_output_locations(self):
        skill_text = (Path(__file__).parents[1] / "SKILL.md").read_text(encoding="utf-8")
        workflow_text = (
            Path(__file__).parents[1] / "references" / "workflow-and-gates.md"
        ).read_text(encoding="utf-8")
        required_rules = [
            "### Required handling of unmatched requirements",
            "after all selected rows for that space and immediately before the space subtotal",
            "`未匹配需求汇总` worksheet MUST receive one row for every unmatched requirement",
            "leave `产品图片` and every unit-price and amount cell blank",
            "excluded from every subtotal, grand total, `confirmed_totals`, and `internal_product_mapping`",
        ]
        for required_rule in required_rules:
            with self.subTest(required_rule=required_rule):
                self.assertIn(required_rule, skill_text)
        self.assertIn("### 未匹配需求的强制输出", workflow_text)
        self.assertIn("已匹配产品行之后、空间小计之前", workflow_text)
        self.assertIn("`未匹配需求汇总`工作表", workflow_text)

    def test_default_template_is_generic_and_blank(self):
        template = Path(
            os.environ.get(
                "DEFAULT_TEMPLATE_PATH",
                Path(__file__).parents[1] / "assets" / "default-configuration-list.xlsx",
            )
        )
        workbook = load_workbook(template, data_only=False)
        self.assertEqual(workbook.sheetnames, ["配置清单", "未匹配需求汇总"])
        sheet = workbook["配置清单"]
        self.assertEqual(sheet["A1"].value, "{{学校名称}}生涯心理学生发展中心配置清单")
        self.assertEqual(sheet["E2"].value, "产品图片")
        self.assertEqual(sheet.freeze_panes, "A3")
        self.assertEqual(sheet.print_area, "'配置清单'!$A$1:$N$2")
        self.assertEqual({str(item) for item in sheet.merged_cells.ranges}, {"A1:N1"})
        self.assertFalse(
            any(
                cell.value is not None
                for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row)
                for cell in row
            )
        )
        self.assertTrue(all(sheet.cell(3, column).has_style for column in range(1, 15)))
        self.assertEqual(len(sheet._images), 0)

        unmatched = workbook["未匹配需求汇总"]
        self.assertEqual(unmatched["A1"].value, "{{学校名称}}生涯心理学生发展中心未匹配需求汇总")
        self.assertEqual(unmatched.freeze_panes, "A3")
        self.assertEqual(unmatched.print_area, "'未匹配需求汇总'!$A$1:$H$2")
        self.assertEqual({str(item) for item in unmatched.merged_cells.ranges}, {"A1:H1"})
        self.assertFalse(
            any(
                cell.value is not None
                for row in unmatched.iter_rows(min_row=3, max_row=unmatched.max_row)
                for cell in row
            )
        )
        self.assertTrue(all(unmatched.cell(3, column).has_style for column in range(1, 9)))

    def test_inserts_exact_catalog_image_without_overwriting_input(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            assets = root / "assets"
            images = assets / "product-images"
            images.mkdir(parents=True)
            image_path = images / "product.png"
            Image.new("RGB", (120, 80), "navy").save(image_path)
            image_hash = hashlib.sha256(image_path.read_bytes()).hexdigest()
            catalog_path = assets / "standard-products.json"
            catalog_path.write_text(
                json.dumps(
                    {
                        "products": [
                            {
                                "product_key": "测试产品|测试品牌|M-1",
                                "image_refs": [
                                    {
                                        "role": "primary",
                                        "path": "product-images/product.png",
                                        "sha256": image_hash,
                                        "confidence": "exact",
                                    }
                                ],
                            }
                        ]
                    }
                ),
                encoding="utf-8",
            )
            input_path = root / "input.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "清单"
            sheet.append(["序号", "品目", "数量"])
            sheet.append([1, "测试产品", 1])
            workbook.save(input_path)
            output_path = root / "output.xlsx"
            result = insert_images(
                input_path=input_path,
                output_path=output_path,
                catalog_path=catalog_path,
                assets_root=assets,
                sheet_name="清单",
                header_row=1,
                bindings=[{"row": 2, "product_key": "测试产品|测试品牌|M-1"}],
            )
            self.assertEqual(result["inserted"], 1)
            self.assertEqual(load_workbook(input_path)["清单"].max_column, 3)
            output_sheet = load_workbook(output_path)["清单"]
            self.assertEqual(output_sheet.cell(1, 4).value, "产品图片")
            self.assertEqual(len(output_sheet._images), 1)


if __name__ == "__main__":
    unittest.main()
