#!/usr/bin/env python3
"""Import WPS/Excel cell images into the configuration-list product catalog.

The source workbook is never changed.  The command writes a new catalog, a
deduplicated display-image directory, and an audit report.  Only rows that
uniquely match a catalog record by normalized name, brand, and model are
linked.  All other source rows are reported as unresolved.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import unicodedata
from collections import defaultdict
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image


WPS_IMAGE_PATTERN = re.compile(r'DISPIMG\("([^"]+)"', re.IGNORECASE)
MATCH_RULE = "normalized-name-brand-model-v1"
IMPORT_SOURCE = "original-product-workbook-v1"
NAMESPACES = {
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def normalize_identity_part(value: object) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).strip()
    text = text.replace("—", "-").replace("–", "-").replace("－", "-")
    return re.sub(r"\s+", "", text)


def identity_key(name: object, brand: object, model: object) -> tuple[str, str, str]:
    return (
        normalize_identity_part(name),
        normalize_identity_part(brand),
        normalize_identity_part(model),
    )


def find_column(headers: list[object], expected: str) -> int:
    expected_value = normalize_identity_part(expected.lstrip("*"))
    for index, header in enumerate(headers, start=1):
        if normalize_identity_part(str(header or "").lstrip("*")) == expected_value:
            return index
    raise ValueError(f"Missing required source column: {expected}")


def select_source_sheet(workbook, requested_name: str | None):
    if requested_name:
        if requested_name not in workbook.sheetnames:
            raise ValueError(f"Requested worksheet was not found: {requested_name}")
        return workbook[requested_name]
    for candidate in ("总清单 (上传版)", "总清单"):
        if candidate in workbook.sheetnames:
            return workbook[candidate]
    raise ValueError("No supported source sheet found. Specify --sheet explicitly.")


def wps_cell_image_targets(workbook_path: Path) -> dict[str, str]:
    with ZipFile(workbook_path) as archive:
        try:
            image_tree = ET.fromstring(archive.read("xl/cellimages.xml"))
            relation_tree = ET.fromstring(archive.read("xl/_rels/cellimages.xml.rels"))
        except KeyError as exc:
            raise ValueError("Source workbook has no WPS cell-image package.") from exc
        target_by_relation = {
            relation.attrib["Id"]: relation.attrib["Target"].replace("../", "")
            for relation in relation_tree.findall("rel:Relationship", NAMESPACES)
        }
        targets: dict[str, str] = {}
        for picture in image_tree.findall(".//xdr:pic", NAMESPACES):
            name = picture.find(".//xdr:cNvPr", NAMESPACES)
            blip = picture.find(".//a:blip", NAMESPACES)
            if name is None or blip is None:
                continue
            relation_id = blip.attrib.get(f"{{{NAMESPACES['r']}}}embed")
            target = target_by_relation.get(relation_id)
            if name.attrib.get("name") and target:
                targets[name.attrib["name"]] = target
        return targets


def source_records(workbook_path: Path, requested_sheet: str | None) -> tuple[str, list[dict], dict[str, str]]:
    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    sheet = select_source_sheet(workbook, requested_sheet)
    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    name_column = find_column(headers, "产品名称")
    brand_column = find_column(headers, "品牌")
    model_column = find_column(headers, "型号")
    image_column = find_column(headers, "产品图片")
    targets = wps_cell_image_targets(workbook_path)
    records = []
    for row in range(2, sheet.max_row + 1):
        name = sheet.cell(row, name_column).value
        brand = sheet.cell(row, brand_column).value
        model = sheet.cell(row, model_column).value
        if not (name and brand and model):
            continue
        image_formula = sheet.cell(row, image_column).value
        match = WPS_IMAGE_PATTERN.search(image_formula) if isinstance(image_formula, str) else None
        image_id = match.group(1) if match else None
        records.append(
            {
                "row": row,
                "name": str(name),
                "brand": str(brand),
                "model": str(model),
                "identity": identity_key(name, brand, model),
                "image_cell": f"{get_column_letter(image_column)}{row}",
                "image_id": image_id,
                "source_media": targets.get(image_id) if image_id else None,
            }
        )
    return sheet.title, records, targets


def display_asset(source_bytes: bytes, maximum_side: int) -> tuple[bytes, str, str]:
    with Image.open(BytesIO(source_bytes)) as image:
        image.load()
        image.thumbnail((maximum_side, maximum_side), Image.Resampling.LANCZOS)
        has_alpha = "A" in image.getbands()
        output = BytesIO()
        if has_alpha:
            image.save(output, format="PNG", optimize=True)
            return output.getvalue(), ".png", "image/png"
        image.convert("RGB").save(output, format="JPEG", quality=85, optimize=True)
        return output.getvalue(), ".jpg", "image/jpeg"


def catalog_index(products: list[dict]) -> dict[tuple[str, str, str], list[dict]]:
    index: dict[tuple[str, str, str], list[dict]] = defaultdict(list)
    for product in products:
        index[identity_key(product.get("product_name"), product.get("brand"), product.get("model"))].append(product)
    return index


def group_source_records(records: list[dict]) -> dict[tuple[str, str, str], list[dict]]:
    grouped: dict[tuple[str, str, str], list[dict]] = defaultdict(list)
    for record in records:
        grouped[record["identity"]].append(record)
    return grouped


def linked_image_ref(record: dict, asset_path: str, source_hash: str, asset_hash: str, mime_type: str) -> dict:
    return {
        "role": "primary",
        "path": asset_path,
        "sha256": asset_hash,
        "mime_type": mime_type,
        "source_sha256": source_hash,
        "import_source": IMPORT_SOURCE,
        "source": {
            "workbook": record["source_workbook"],
            "worksheet": record["source_worksheet"],
            "cell": record["image_cell"],
            "cell_image_id": record["image_id"],
        },
        "match_rule": MATCH_RULE,
        "confidence": "exact",
    }


def update_managed_image_refs(product: dict, image_ref: dict) -> None:
    existing = product.get("image_refs", [])
    product["image_refs"] = [
        item for item in existing if item.get("import_source") != IMPORT_SOURCE
    ] + [image_ref]


def import_images(args: argparse.Namespace) -> dict:
    source_workbook = args.source_xlsx.resolve()
    output_catalog = args.output_catalog.resolve()
    images_directory = args.images_directory.resolve()
    report_path = args.report.resolve()
    catalog = json.loads(args.catalog.read_text(encoding="utf-8"))
    products = catalog.get("products")
    if not isinstance(products, list):
        raise ValueError("Catalog must contain a products array.")
    sheet_name, rows, cell_image_targets = source_records(source_workbook, args.sheet)
    source_groups = group_source_records(rows)
    product_groups = catalog_index(products)
    source_hash = sha256_file(source_workbook)
    successful_links = []
    unresolved = []
    asset_cache: dict[str, dict] = {}
    images_directory.mkdir(parents=True, exist_ok=True)

    with ZipFile(source_workbook) as archive:
        for identity, source_group in sorted(source_groups.items()):
            image_ids = {record["image_id"] for record in source_group if record["image_id"]}
            media_paths = {record["source_media"] for record in source_group if record["source_media"]}
            candidates = product_groups.get(identity, [])
            representative = source_group[0]
            if not image_ids:
                unresolved.append({"reason": "no_source_image", "source_rows": source_group})
                continue
            if len(image_ids) != 1 or len(media_paths) != 1:
                unresolved.append({"reason": "conflicting_source_images", "source_rows": source_group})
                continue
            if len(candidates) != 1:
                unresolved.append(
                    {
                        "reason": "missing_catalog_identity" if not candidates else "non_unique_catalog_identity",
                        "source_rows": source_group,
                        "catalog_candidates": [product.get("product_key") for product in candidates],
                    }
                )
                continue
            media_path = next(iter(media_paths))
            try:
                original_bytes = archive.read(f"xl/{media_path}")
            except KeyError:
                unresolved.append({"reason": "missing_source_media", "source_rows": source_group})
                continue
            original_hash = sha256_bytes(original_bytes)
            if original_hash not in asset_cache:
                asset_bytes, extension, mime_type = display_asset(original_bytes, args.maximum_side)
                asset_hash = sha256_bytes(asset_bytes)
                relative_path = f"product-images/{asset_hash}{extension}"
                destination = images_directory / f"{asset_hash}{extension}"
                if not destination.exists():
                    destination.write_bytes(asset_bytes)
                asset_cache[original_hash] = {
                    "relative_path": relative_path,
                    "asset_hash": asset_hash,
                    "mime_type": mime_type,
                    "source_hash": original_hash,
                }
            asset = asset_cache[original_hash]
            representative.update(
                {"source_workbook": source_workbook.name, "source_worksheet": sheet_name}
            )
            product = candidates[0]
            image_ref = linked_image_ref(
                representative,
                asset["relative_path"],
                asset["source_hash"],
                asset["asset_hash"],
                asset["mime_type"],
            )
            update_managed_image_refs(product, image_ref)
            successful_links.append(
                {
                    "product_key": product.get("product_key"),
                    "product_identity_key": product.get("product_identity_key"),
                    "source_rows": [record["row"] for record in source_group],
                    "image_ref": image_ref,
                }
            )

    catalog["image_catalog"] = {
        "schema_version": "1.0",
        "asset_root": "product-images",
        "display_maximum_side_px": args.maximum_side,
        "source_workbook": {
            "name": source_workbook.name,
            "sha256": source_hash,
            "worksheet": sheet_name,
            "image_column": "产品图片",
        },
        "match_rule": MATCH_RULE,
        "report": "image-import-report.json",
    }
    report = {
        "schema_version": "1.0",
        "source_workbook": catalog["image_catalog"]["source_workbook"],
        "match_rule": MATCH_RULE,
        "summary": {
            "source_rows": len(rows),
            "unique_source_identities": len(source_groups),
            "source_cell_image_identifiers": len(cell_image_targets),
            "linked_catalog_products": len(successful_links),
            "unique_display_assets": len(asset_cache),
            "unresolved_source_identities": len(unresolved),
        },
        "links": successful_links,
        "unresolved": unresolved,
    }
    output_catalog.parent.mkdir(parents=True, exist_ok=True)
    output_catalog.write_text(json.dumps(catalog, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return report


def arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source-xlsx", required=True, type=Path)
    parser.add_argument("--catalog", required=True, type=Path, help="Read-only input catalog")
    parser.add_argument("--output-catalog", required=True, type=Path)
    parser.add_argument("--images-directory", required=True, type=Path)
    parser.add_argument("--report", required=True, type=Path)
    parser.add_argument("--sheet", help="Default: 总清单 (上传版), then 总清单")
    parser.add_argument("--maximum-side", type=int, default=640)
    return parser.parse_args()


def main() -> None:
    args = arguments()
    if args.maximum_side < 64:
        raise ValueError("--maximum-side must be at least 64 pixels.")
    report = import_images(args)
    print(json.dumps(report["summary"], ensure_ascii=False))


if __name__ == "__main__":
    main()
