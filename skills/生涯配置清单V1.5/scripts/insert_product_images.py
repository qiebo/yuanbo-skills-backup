#!/usr/bin/env python3
"""Insert verified product thumbnails into a newly generated configuration list.

The binding manifest must identify each output row by the catalog's unique
product_key.  This intentionally refuses product-name-only matching.
"""

from __future__ import annotations

import argparse
import hashlib
import json
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from PIL import Image


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def safe_asset_path(assets_root: Path, relative_path: str) -> Path:
    root = assets_root.resolve()
    candidate = (root / relative_path).resolve()
    try:
        candidate.relative_to(root)
    except ValueError as exc:
        raise ValueError(f"Image path escapes the asset directory: {relative_path}") from exc
    return candidate


def exact_primary_image(product: dict, assets_root: Path) -> Path:
    references = [
        reference
        for reference in product.get("image_refs", [])
        if reference.get("role") == "primary" and reference.get("confidence") == "exact"
    ]
    if len(references) != 1:
        raise ValueError("Product has no single, exact primary image reference.")
    reference = references[0]
    path = safe_asset_path(assets_root, reference.get("path", ""))
    if not path.is_file():
        raise ValueError(f"Referenced image file is missing: {reference.get('path')}")
    if sha256_file(path) != reference.get("sha256"):
        raise ValueError(f"Referenced image file failed SHA-256 verification: {reference.get('path')}")
    return path


def image_column(sheet, header_row: int, header: str = "产品图片") -> int:
    for column in range(1, sheet.max_column + 1):
        if sheet.cell(header_row, column).value == header:
            return column
    column = sheet.max_column + 1
    anchor = sheet.cell(header_row, max(1, column - 1))
    target = sheet.cell(header_row, column, header)
    if anchor.has_style:
        target._style = copy(anchor._style)
        target.font = copy(anchor.font)
        target.fill = copy(anchor.fill)
        target.border = copy(anchor.border)
        target.number_format = anchor.number_format
        target.protection = copy(anchor.protection)
    target.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.column_dimensions[get_column_letter(column)].width = 17
    return column


def fit_image(path: Path, maximum_width: int = 100, maximum_height: int = 90) -> ExcelImage:
    with Image.open(path) as source:
        width, height = source.size
    scale = min(maximum_width / width, maximum_height / height, 1)
    image = ExcelImage(str(path))
    image.width = max(1, round(width * scale))
    image.height = max(1, round(height * scale))
    return image


def insert_images(
    *,
    input_path: Path,
    output_path: Path,
    catalog_path: Path,
    assets_root: Path,
    sheet_name: str,
    header_row: int,
    bindings: list[dict],
) -> dict:
    input_path = input_path.resolve()
    output_path = output_path.resolve()
    if input_path == output_path:
        raise ValueError("Input and output workbooks must be different files.")
    catalog = json.loads(catalog_path.read_text(encoding="utf-8"))
    products_by_key = {}
    for product in catalog.get("products", []):
        product_key = product.get("product_key")
        if not product_key:
            continue
        if product_key in products_by_key:
            raise ValueError(f"Duplicate product_key in catalog: {product_key}")
        products_by_key[product_key] = product
    workbook = load_workbook(input_path)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Output worksheet was not found: {sheet_name}")
    sheet = workbook[sheet_name]
    column = image_column(sheet, header_row)
    seen_rows = set()
    warnings = []
    inserted = 0
    for binding in bindings:
        row = binding.get("row")
        product_key = binding.get("product_key")
        if not isinstance(row, int) or row <= header_row:
            raise ValueError(f"Invalid output row in binding: {row}")
        if row in seen_rows:
            raise ValueError(f"Duplicate output row in binding manifest: {row}")
        seen_rows.add(row)
        product = products_by_key.get(product_key)
        if product is None:
            raise ValueError(f"Binding references an unknown product_key: {product_key}")
        try:
            asset = exact_primary_image(product, assets_root)
        except ValueError as exc:
            warnings.append({"row": row, "product_key": product_key, "reason": str(exc)})
            continue
        image = fit_image(asset)
        sheet.add_image(image, f"{get_column_letter(column)}{row}")
        current_height = sheet.row_dimensions[row].height or 15
        sheet.row_dimensions[row].height = max(current_height, image.height * 0.75 + 6)
        inserted += 1
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return {"inserted": inserted, "warnings": warnings, "image_column": get_column_letter(column)}


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input-xlsx", required=True, type=Path)
    parser.add_argument("--output-xlsx", required=True, type=Path)
    parser.add_argument("--catalog", required=True, type=Path)
    parser.add_argument("--assets-root", required=True, type=Path)
    parser.add_argument("--manifest", required=True, type=Path)
    args = parser.parse_args()
    manifest = json.loads(args.manifest.read_text(encoding="utf-8"))
    result = insert_images(
        input_path=args.input_xlsx,
        output_path=args.output_xlsx,
        catalog_path=args.catalog,
        assets_root=args.assets_root,
        sheet_name=manifest["sheet_name"],
        header_row=manifest["header_row"],
        bindings=manifest["bindings"],
    )
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
